"""
Report Service — PDF & Excel report generation
================================================

Generates professional PDF (reportlab) and Excel (openpyxl) exports of
system data: attendance registers, employee stats, recognition logs,
unknown-face reports, and audit logs.

Usage::

    from services.report_service import ReportService

    rows = ReportService.attendance_rows(date.today())
    pdf_bytes = ReportService.attendance_pdf(rows, title="Daily Attendance")
    xlsx_bytes = ReportService.attendance_excel(rows, title="Daily Attendance")

Design notes:
- The heavy libraries (reportlab / openpyxl) are imported lazily inside the
  generator functions so the rest of the app works fine when they are not
  installed. Missing dependencies raise ``ReportUnavailableError`` with a
  clear install hint — callers should catch it and show a friendly message.
- Rows are plain ``list[dict]`` so the same generator works for attendance,
  audit logs, unknown faces, etc.
- Cell values are escaped/truncated defensively; nothing here ever touches
  biometric data (embeddings, face crops).
"""

from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence

from sqlalchemy.orm import selectinload

from database.database import get_session
from database.models import Attendance, Employee
from database.repository import AttendanceRepo, AuditLogRepo

logger = logging.getLogger(__name__)


class ReportUnavailableError(RuntimeError):
    """Raised when a report cannot be generated (missing dependency)."""


# ── Data helpers ─────────────────────────────────────────────────


class ReportService:
    """Static report generation helpers."""

    # ── Row builders (plain dicts, safe for PDF/Excel) ─────────

    @staticmethod
    def attendance_rows(
        target_date: date,
        limit: Optional[int] = None,
        skip: int = 0,
    ) -> List[Dict[str, Any]]:
        """Fetch attendance for a date as plain dict rows."""
        rows: List[Dict[str, Any]] = []
        with get_session() as session:
            records = AttendanceRepo.get_by_date(
                session,
                target_date,
                limit=limit,
                skip=skip,
            )
            for r in records:
                emp = r.employee
                rows.append(
                    {
                        "ID": emp.employee_id if emp else f"ID:{r.employee_id}",
                        "Name": emp.name if emp else "Unknown",
                        "Department": emp.department if emp and emp.department else "—",
                        "Time": r.timestamp.strftime("%Y-%m-%d %H:%M:%S") if r.timestamp else "",
                        "Confidence": f"{r.confidence:.1%}" if r.confidence is not None else "",
                    }
                )
        return rows

    @staticmethod
    def attendance_range_rows(
        start: date,
        end: date,
        limit: int = 5000,
    ) -> List[Dict[str, Any]]:
        """Fetch attendance across a date range (bounded) as plain dict rows."""
        rows: List[Dict[str, Any]] = []
        start_dt = datetime.combine(start, datetime.min.time())
        end_dt = datetime.combine(end, datetime.max.time())
        with get_session() as session:
            records = (
                session.query(Attendance)
                .filter(Attendance.timestamp.between(start_dt, end_dt))
                .options(selectinload(Attendance.employee))
                .order_by(Attendance.timestamp.desc())
                .limit(limit)
                .all()
            )
            for r in records:
                emp = r.employee
                rows.append(
                    {
                        "Date": r.timestamp.strftime("%Y-%m-%d") if r.timestamp else "",
                        "ID": emp.employee_id if emp else f"ID:{r.employee_id}",
                        "Name": emp.name if emp else "Unknown",
                        "Department": emp.department if emp and emp.department else "—",
                        "Time": r.timestamp.strftime("%H:%M:%S") if r.timestamp else "",
                        "Confidence": f"{r.confidence:.1%}" if r.confidence is not None else "",
                    }
                )
        return rows

    @staticmethod
    def audit_rows(
        query: str = "",
        action: Optional[str] = None,
        severity: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        limit: int = 2000,
        skip: int = 0,
    ) -> List[Dict[str, Any]]:
        """Fetch audit logs as plain dict rows."""
        rows: List[Dict[str, Any]] = []
        with get_session() as session:
            page = AuditLogRepo.search_paginated(
                session,
                query=query,
                action=action,
                severity=severity,
                date_from=date_from,
                date_to=date_to,
                limit=limit,
                skip=skip,
            )
            for log in page.items:
                rows.append(
                    {
                        "Timestamp": log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else "",
                        "Action": log.action or "",
                        "Actor": log.actor or "",
                        "Severity": log.severity or "INFO",
                        "Resource": log.resource_type or "",
                        "Description": log.description or "",
                        "IP": log.ip_address or "",
                    }
                )
        return rows

    @staticmethod
    def employee_rows(limit: int = 2000) -> List[Dict[str, Any]]:
        """Fetch employees as plain dict rows."""
        rows: List[Dict[str, Any]] = []
        with get_session() as session:
            employees = session.query(Employee).order_by(Employee.name, Employee.id).limit(limit).all()
            for e in employees:
                rows.append(
                    {
                        "ID": e.employee_id or "",
                        "Name": e.name or "",
                        "Department": e.department or "—",
                    }
                )
        return rows

    # ── Generic table rendering (PDF / Excel) ─────────────────

    @staticmethod
    def _pdf_table(
        title: str,
        columns: Sequence[str],
        rows: Sequence[Dict[str, Any]],
        subtitle: str = "",
    ) -> bytes:
        """Render a titled table to PDF bytes (reportlab)."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError as exc:  # pragma: no cover
            raise ReportUnavailableError(
                "PDF export requires 'reportlab'. Install with: pip install reportlab"
            ) from exc

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            fontSize=16,
            spaceAfter=4,
        )
        subtitle_style = ParagraphStyle(
            "ReportSubtitle",
            parent=styles["Normal"],
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=8,
        )
        header_style = ParagraphStyle(
            "ReportHeader",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.white,
            fontName="Helvetica-Bold",
        )
        cell_style = ParagraphStyle(
            "ReportCell",
            parent=styles["Normal"],
            fontSize=7.5,
            leading=9,
        )

        story = [Paragraph(_escape(title), title_style)]
        if subtitle:
            story.append(Paragraph(_escape(subtitle), subtitle_style))
        story.append(Spacer(1, 6))

        def _fmt(value: Any) -> str:
            if value is None:
                return ""
            text = str(value)
            return text if len(text) <= 120 else text[:117] + "…"

        table_data: List[List[Any]] = [
            [Paragraph(_escape(c), header_style) for c in columns],
        ]
        for row in rows[:500]:  # Hard cap: 500 rows per PDF page-set
            table_data.append([Paragraph(_escape(_fmt(row.get(c, ""))), cell_style) for c in columns])

        table = Table(table_data, repeatRows=1, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)

        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def _excel_table(
        title: str,
        columns: Sequence[str],
        rows: Sequence[Dict[str, Any]],
        sheet_name: str = "Report",
    ) -> bytes:
        """Render a titled table to XLSX bytes (openpyxl)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover
            raise ReportUnavailableError(
                "Excel export requires 'openpyxl'. Install with: pip install openpyxl"
            ) from exc

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title row
        ws.append([title])
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))

        # Header row
        ws.append(list(columns))
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row in rows[:5000]:  # Hard cap: 5000 rows per workbook
            ws.append([row.get(c, "") for c in columns])

        # Column widths + freeze header
        for i, col in enumerate(columns, start=1):
            col_letter = get_column_letter(i)
            width = max(len(str(col)) + 2, 10)
            for j in range(min(len(rows), 100) + 2):
                value = ws.cell(row=j + 1, column=i).value
                if value is not None:
                    width = max(width, min(len(str(value)) + 2, 60))
            ws.column_dimensions[col_letter].width = width
        ws.freeze_panes = "A3"

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ── High-level report generators ──────────────────────────

    @staticmethod
    def attendance_pdf(
        target_date: date,
        title: Optional[str] = None,
        rows: Optional[List[Dict[str, Any]]] = None,
    ) -> bytes:
        """PDF attendance register for a date."""
        rows = rows if rows is not None else ReportService.attendance_rows(target_date)
        title = title or f"Attendance Register — {target_date.strftime('%d %b %Y')}"
        subtitle = f"{len(rows)} record(s) · Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        return ReportService._pdf_table(
            title,
            ["ID", "Name", "Department", "Time", "Confidence"],
            rows,
            subtitle,
        )

    @staticmethod
    def attendance_excel(
        target_date: date,
        title: Optional[str] = None,
        rows: Optional[List[Dict[str, Any]]] = None,
    ) -> bytes:
        """Excel attendance register for a date."""
        rows = rows if rows is not None else ReportService.attendance_rows(target_date)
        title = title or f"Attendance Register — {target_date.strftime('%d %b %Y')}"
        return ReportService._excel_table(
            title,
            ["ID", "Name", "Department", "Time", "Confidence"],
            rows,
            sheet_name="Attendance",
        )

    @staticmethod
    def audit_pdf(
        query: str = "",
        action: Optional[str] = None,
        severity: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        title: Optional[str] = None,
    ) -> bytes:
        """PDF audit log export with optional filters."""
        rows = ReportService.audit_rows(
            query=query,
            action=action,
            severity=severity,
            date_from=date_from,
            date_to=date_to,
        )
        title = title or "Audit Log Export"
        subtitle = f"{len(rows)} entrie(s) · Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        return ReportService._pdf_table(
            title,
            ["Timestamp", "Action", "Actor", "Severity", "Resource", "Description", "IP"],
            rows,
            subtitle,
        )

    @staticmethod
    def audit_excel(
        query: str = "",
        action: Optional[str] = None,
        severity: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        title: Optional[str] = None,
    ) -> bytes:
        """Excel audit log export with optional filters."""
        rows = ReportService.audit_rows(
            query=query,
            action=action,
            severity=severity,
            date_from=date_from,
            date_to=date_to,
        )
        title = title or "Audit Log Export"
        return ReportService._excel_table(
            title,
            ["Timestamp", "Action", "Actor", "Severity", "Resource", "Description", "IP"],
            rows,
            sheet_name="AuditLog",
        )

    @staticmethod
    def employees_excel(title: Optional[str] = None) -> bytes:
        """Excel employee directory export."""
        rows = ReportService.employee_rows()
        title = title or "Employee Directory"
        return ReportService._excel_table(
            title,
            ["ID", "Name", "Department"],
            rows,
            sheet_name="Employees",
        )

    @staticmethod
    def employees_pdf(title: Optional[str] = None) -> bytes:
        """PDF employee directory export."""
        rows = ReportService.employee_rows()
        title = title or "Employee Directory"
        subtitle = f"{len(rows)} employee(s) · Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        return ReportService._pdf_table(
            title,
            ["ID", "Name", "Department"],
            rows,
            subtitle,
        )


def _escape(text: Any) -> str:
    """Escape text for reportlab Paragraph markup."""
    if text is None:
        return ""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
